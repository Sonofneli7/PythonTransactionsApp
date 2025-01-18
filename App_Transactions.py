import openpyxl as xl       # giving openpyxl an alias 'xl'
from openpyxl.chart import BarChart, Reference  # in library pkg have module chart and 2 classes arChart, Reference


def process_workbook(filename):     # defining a function to automate 1000's of spreadsheets
    #tabbed all code underneath to put inside the function
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']    # accessing sheet in workbook

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)        # getting access to row objects in column 3
        corrected_price = cell.value * 0.9      # mistake going to fix with program
        corrected_price_cell = sheet.cell(row, 4)      # writing corrected price direct to sheet
        corrected_price_cell.value = corrected_price          # setting value of corrected cell to correct price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)       # using Reference class to select a range of values & only need values in 4th column, store in values object

    # half of prb solved # now add chart
    chart = BarChart()         # Creating instance of BarChart() class and store in 'chart' object
    chart.add_data(values)     # call the chart & add data with underlying values
    sheet.add_chart(chart, 'e2')                           # call the sheet and add chart passing in the 'chart' object, specify where to add chart (put in cell coordinates)

    # Saving new information in overridden transactions.xlsx
    wb.save(filename)


